from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from pathlib import Path
import csv
from datetime import datetime
import uuid

app = Flask(__name__)
CORS(app)


BASE_DIR = Path(__file__).resolve().parent

CSV_FILE = BASE_DIR / "registrations.csv"
EXCEL_FILE = BASE_DIR / "registrations.xlsx"


HEADERS = [
    "Registration ID",
    "Name",
    "Email",
    "Phone",
    "College",
    "Course",
    "Year",
    "Skills",
    "GitHub",
    "LinkedIn",
    "Registered At"
]


def create_csv_if_not_exists():

    if not CSV_FILE.exists():

        with open(
            CSV_FILE,
            "w",
            newline="",
            encoding="utf-8"
        ) as file:

            writer = csv.writer(file)

            writer.writerow(HEADERS)


# =========================================================
# CREATE EXCEL FILE
# =========================================================

def create_excel_if_not_exists():

    if not EXCEL_FILE.exists():

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Registrations"

        sheet.append(HEADERS)

        # Make columns readable
        widths = [
            20, 25, 30, 18, 35,
            25, 15, 40, 40, 40, 25
        ]

        for index, width in enumerate(widths, start=1):

            sheet.column_dimensions[
                chr(64 + index)
            ].width = width

        workbook.save(EXCEL_FILE)


# =========================================================
# SAVE REGISTRATION TO CSV
# =========================================================

def save_to_csv(data):

    create_csv_if_not_exists()

    with open(
        CSV_FILE,
        "a",
        newline="",
        encoding="utf-8"
    ) as file:

        writer = csv.writer(file)

        writer.writerow([
            data["id"],
            data["name"],
            data["email"],
            data["phone"],
            data["college"],
            data["course"],
            data["year"],
            data["skills"],
            data["github"],
            data["linkedin"],
            data["registeredAt"]
        ])


# =========================================================
# SAVE REGISTRATION TO EXCEL
# =========================================================

def save_to_excel(data):

    create_excel_if_not_exists()

    workbook = load_workbook(EXCEL_FILE)

    sheet = workbook["Registrations"]

    sheet.append([
        data["id"],
        data["name"],
        data["email"],
        data["phone"],
        data["college"],
        data["course"],
        data["year"],
        data["skills"],
        data["github"],
        data["linkedin"],
        data["registeredAt"]
    ])

    workbook.save(EXCEL_FILE)


# =========================================================
# REGISTRATION API
# =========================================================

@app.route("/api/register", methods=["POST"])
def register():

    try:

        data = request.get_json()

        if not data:

            return jsonify({
                "success": False,
                "message": "No registration data received."
            }), 400


        # Required fields
        required_fields = [
            "name",
            "email",
            "phone",
            "college",
            "course",
            "year"
        ]


        for field in required_fields:

            if not str(data.get(field, "")).strip():

                return jsonify({
                    "success": False,
                    "message": f"{field} is required."
                }), 400


        # Generate registration ID
        registration_id = (
            "CC-"
            + datetime.now().strftime("%Y%m%d")
            + "-"
            + uuid.uuid4().hex[:6].upper()
        )


        registration = {

            "id": registration_id,

            "name":
                str(data.get("name", "")).strip(),

            "email":
                str(data.get("email", "")).strip(),

            "phone":
                str(data.get("phone", "")).strip(),

            "college":
                str(data.get("college", "")).strip(),

            "course":
                str(data.get("course", "")).strip(),

            "year":
                str(data.get("year", "")).strip(),

            "skills":
                str(data.get("skills", "")).strip(),

            "github":
                str(data.get("github", "")).strip(),

            "linkedin":
                str(data.get("linkedin", "")).strip(),

            "registeredAt":
                datetime.now().isoformat()

        }


        # Save to both files
        save_to_csv(registration)

        save_to_excel(registration)


        return jsonify({

            "success": True,

            "message":
                "Registration successful!",

            "registration": registration

        }), 201


    except Exception as error:

        print("ERROR:", error)

        return jsonify({

            "success": False,

            "message":
                "Unable to save registration."

        }), 500


# =========================================================
# GET ALL REGISTRATIONS
# =========================================================

@app.route("/api/registrations", methods=["GET"])
def get_registrations():

    try:

        create_csv_if_not_exists()

        registrations = []

        with open(
            CSV_FILE,
            "r",
            encoding="utf-8"
        ) as file:

            reader = csv.DictReader(file)

            for row in reader:

                registrations.append(row)


        return jsonify({

            "success": True,

            "registrations":
                registrations

        })


    except Exception as error:

        print("ERROR:", error)

        return jsonify({

            "success": False,

            "message":
                "Unable to load registrations."

        }), 500


# =========================================================
# SERVER HEALTH CHECK
# =========================================================

@app.route("/", methods=["GET"])
def home():

    return jsonify({

        "project":
            "CloudConnect",

        "status":
            "Backend is running",

        "message":
            "CloudConnect Registration API"

    })


# =========================================================
# START SERVER
# =========================================================

if __name__ == "__main__":

    create_csv_if_not_exists()

    create_excel_if_not_exists()

    print("-----------------------------------------")
    print("CloudConnect Backend")
    print("-----------------------------------------")
    print("Server: http://127.0.0.1:5000")
    print("CSV:", CSV_FILE)
    print("Excel:", EXCEL_FILE)
    print("-----------------------------------------")

    app.run(
        host="127.0.0.1",
        port=5000,
        debug=True
    )

